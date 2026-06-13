"""
generate_rnd_excel.py
=====================
Generates a formatted Excel R&D report.
Run: python generate_rnd_excel.py
"""

import pandas as pd
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "RnD_Phase2_Report.xlsx")

# ── Sheet 1: Tool Research Summary ──────────────────────────────────────────
tools_data = {
    "Tool": [
        "Turnitin",
        "GPTZero",
        "Originality.ai",
        "Copyleaks"
    ],
    "What It Does": [
        "Breaks documents into overlapping 5-word phrase chunks, converts each into a unique numeric fingerprint, compares fingerprints across all submissions. Detects even partial paragraph copying.",
        "Measures two signals: Perplexity (how predictable each word choice is — AI picks safe words, humans pick unexpected ones) and Burstiness (AI has uniform sentence lengths, humans mix short and long).",
        "Runs each submission through 5+ independent AI detection models simultaneously and takes majority vote. Also measures Flesch-Kincaid Grade Level — AI always writes at PhD level, students do not.",
        "Reads hidden metadata inside Word/PDF files — creation timestamp, last-saved timestamp, author name. Calculates whether human typing speed could have produced that word count in that time."
    ],
    "Core Algorithm": [
        "Shingling (n-gram extraction) + Hash function → Jaccard similarity coefficient between fingerprint sets",
        "Perplexity: language model probability scoring per token. Burstiness: standard deviation of sentence length distribution",
        "Ensemble model aggregation (weighted average across 5+ models) + Flesch-Kincaid formula: GL = (0.39 x ASL) + (11.8 x ASW) - 15.59",
        "File metadata parsing (DOCX XML core properties) + Typing speed calculation: min_time = word_count / 30 wpm"
    ],
    "Key Strength": [
        "Catches partial copying that whole-document comparison misses",
        "Mathematical signal — not affected by writing style. Catches AI even when it sounds casual",
        "Lowest false positive rate in industry. Grade level mismatch catches AI that passed style detection",
        "Non-text forensic signal. Cannot be defeated by editing the content"
    ],
    "Gap in Our Current System": [
        "TF-IDF compares whole documents. Student copying 3 paragraphs + writing rest scores low and passes",
        "AI detector judges style — wrongly penalises students who write clearly and formally (false positives)",
        "One detector firing high drags down integrity score. Grade level never measured",
        "Only text content is analysed. File timestamps not checked at all"
    ]
}

# ── Sheet 2: Proposed Features ───────────────────────────────────────────────
features_data = {
    "Feature": [
        "N-Gram Phrase Fingerprinting",
        "Perplexity Scoring",
        "Grade Level Check + Majority Voting",
        "Metadata Timing Forensics"
    ],
    "Inspired By": [
        "Turnitin",
        "GPTZero",
        "Originality.ai",
        "Copyleaks"
    ],
    "Detection Principle": [
        "Extract all 5-word phrase chunks from each submission. Convert to hash fingerprints. Compare fingerprint sets using Jaccard similarity. Reconstruct matched phrases for teacher evidence.",
        "For each sentence, use existing Ollama model to estimate how predictable the word choices are. Average across document = perplexity score. Low = AI, High = Human.",
        "Apply Flesch-Kincaid formula to measure writing grade level. Flag submissions above Grade 12 for Class 10 students. Count how many detectors agree — require majority before high-confidence flag.",
        "Read Word/PDF file metadata (python-docx already installed). Compare creation-to-save elapsed time against minimum human typing time (word count / 30 wpm). Flag impossibilities."
    ],
    "Gap It Closes": [
        "Partial paragraph copying currently passes TF-IDF whole-document comparison",
        "Genuine students who write clearly are wrongly flagged by style-based detection",
        "Single detector errors cause false positives. AI writing at casual tone passes content checks",
        "AI-generated content pasted into Word file passes all text-based detectors"
    ],
    "Files to Modify": [
        "copy_checker.py, teacher_ui.py",
        "New detector module, report_builder.py, teacher_ui.py",
        "originality.py, report_builder.py, teacher_ui.py",
        "text_extractor.py, report_builder.py, teacher_ui.py"
    ],
    "New Libraries Required": [
        "None",
        "None — uses Ollama already running",
        "None — pure Python formula",
        "None — python-docx already installed"
    ],
    "Expected Impact": [
        "High — catches the most common real-world plagiarism pattern",
        "Very High — directly resolves false positive problem",
        "High — removes false positives, adds independent AI signal",
        "High — detection layer fully independent of text content"
    ]
}

# ── Sheet 3: Current vs Phase 2 ──────────────────────────────────────────────
comparison_data = {
    "Detection Capability": [
        "AI writing detection",
        "Peer copy detection",
        "Paraphrase detection",
        "Originality scoring",
        "Partial paragraph copy detection",
        "Word predictability (perplexity)",
        "Writing grade level check",
        "Multi-detector confidence voting",
        "File metadata forensics"
    ],
    "Current System (Phase 1)": [
        "YES — style-based via Qwen2.5",
        "YES — TF-IDF whole document",
        "YES — embedding similarity",
        "YES — TTR + burstiness + personal markers",
        "NO",
        "NO",
        "NO",
        "NO",
        "NO"
    ],
    "After Phase 2": [
        "YES — style + perplexity scoring",
        "YES — TF-IDF + n-gram fingerprinting",
        "YES — embedding similarity",
        "YES — TTR + burstiness + grade level",
        "YES — n-gram fingerprinting",
        "YES — perplexity via Ollama",
        "YES — Flesch-Kincaid formula",
        "YES — majority voting layer",
        "YES — timestamp forensics"
    ]
}

# ── Write to Excel ────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    pd.DataFrame(tools_data).to_excel(writer, sheet_name="Tool Research", index=False)
    pd.DataFrame(features_data).to_excel(writer, sheet_name="Proposed Features", index=False)
    pd.DataFrame(comparison_data).to_excel(writer, sheet_name="Phase 1 vs Phase 2", index=False)

    # Auto-fit column widths on all sheets
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    # wrap at 80 chars for readability
                    lines = str(cell.value).split("\n")
                    max_len = max(max_len, min(max(len(l) for l in lines), 80))
            ws.column_dimensions[col_letter].width = max_len + 4

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F3864")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Wrap text in all data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[1].height = 30

print(f"\n Excel report saved to:\n  {OUTPUT}\n")
