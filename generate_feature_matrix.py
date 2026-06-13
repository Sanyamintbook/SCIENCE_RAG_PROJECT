"""
generate_feature_matrix.py
===========================
Generates a 4-sheet Excel file.
One sheet per tool. Each sheet lists features of that tool,
whether our project has it (GREEN=YES, RED=NO),
and if NO — whether it can be implemented and why/why not.

Run: python generate_feature_matrix.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "RnD_Feature_Matrix.xlsx"
)

# ── Colour palette ────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill("solid", fgColor="1E8449")   # dark green  → YES
RED_FILL    = PatternFill("solid", fgColor="C0392B")   # dark red    → NO
AMBER_FILL  = PatternFill("solid", fgColor="D68910")   # amber       → PARTIAL
HEADER_FILL = PatternFill("solid", fgColor="1A252F")   # near-black  → header row
TOOL_FILL   = PatternFill("solid", fgColor="2E86C1")   # blue        → tool name row

WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT   = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)

WRAP   = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, tool_name, columns):
    # Row 1 — tool name banner
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1, value=tool_name)
    cell.fill = TOOL_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Row 2 — column headers
    for col, name in enumerate(columns, 1):
        c = ws.cell(row=2, column=col, value=name)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = CENTER
        c.border = THIN
    ws.row_dimensions[2].height = 30


def write_row(ws, row_num, values, status):
    """
    status: 'YES' | 'NO' | 'PARTIAL'
    values: list matching the column count
    The 3rd column (In Our Project) gets the colour fill.
    """
    fill_map = {"YES": GREEN_FILL, "NO": RED_FILL, "PARTIAL": AMBER_FILL}
    text_map = {"YES": "YES ✓", "NO": "NO ✗", "PARTIAL": "PARTIAL ~"}

    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.border = THIN
        c.alignment = WRAP
        c.font = NORMAL_FONT
        if col == 3:                          # "In Our Project" column
            c.fill = fill_map.get(status, PatternFill())
            c.font = Font(color="FFFFFF", bold=True, size=10)
            c.value = text_map.get(status, val)
            c.alignment = CENTER
    ws.row_dimensions[row_num].height = 55


# ═════════════════════════════════════════════════════════════════════════════
# DATA — only school-relevant features (no internet DB, no millions of records)
# Columns: Feature | What It Does | In Our Project | Can Implement | Reason
# ═════════════════════════════════════════════════════════════════════════════

COLS = [
    "Feature",
    "What It Does",
    "In Our Project",
    "Can We Implement It?",
    "Reason / Note"
]

TURNITIN = [
    # (feature, what it does, status, can implement, reason)
    (
        "Peer-to-Peer Copy Detection",
        "Compares each student's submission against every other student in the batch and gives a similarity percentage.",
        "YES", "-", "Already implemented using TF-IDF cosine similarity in copy_checker.py"
    ),
    (
        "AI Writing Detection",
        "Detects whether the submitted text was written by an AI tool like ChatGPT or Gemini.",
        "YES", "-", "Already implemented using Qwen2.5 via Ollama in ai_detector.py"
    ),
    (
        "Paraphrase Detection",
        "Detects when a student rewrote another student's work in different words but same meaning.",
        "YES", "-", "Already implemented using nomic-embed-text semantic embeddings in paraphrase.py"
    ),
    (
        "Combined Integrity Score",
        "Combines all detection signals into a single score and gives CLEAN / REVIEW / FLAGGED verdict.",
        "YES", "-", "Already implemented in report_builder.py"
    ),
    (
        "N-Gram Phrase Fingerprinting",
        "Breaks text into 5-word chunks and compares chunks across submissions. Catches partial paragraph copying that whole-document comparison misses.",
        "NO", "YES — Confidently",
        "Pure Python. No new library needed. Detects copying of 2-3 paragraphs even when the rest of the document is original. Current TF-IDF misses this pattern."
    ),
    (
        "Sentence-Level Peer Match Highlighting",
        "Shows teacher exactly which sentences from Student A matched which sentences from Student B — side by side evidence, not just a percentage.",
        "NO", "YES — Confidently",
        "Extends existing copy_checker.py logic. No new library needed. Gives teacher direct textual evidence instead of an abstract number."
    ),
    (
        "Internet-Wide Source Database",
        "Checks submission against billions of web pages, academic journals, and globally submitted student papers.",
        "NO", "NO",
        "Requires crawling and indexing the entire internet. Needs massive server infrastructure. Not feasible for a local offline system. Paid tools like Turnitin and Copyleaks provide this."
    ),
]

GPTZERO = [
    (
        "AI Probability Score",
        "Gives a 0-100% probability that the submitted text was written by an AI.",
        "YES", "-", "Already implemented in ai_detector.py using Qwen2.5 prompt-based detection."
    ),
    (
        "Sentence-Level AI Flagging",
        "Highlights which specific sentences in the submission appear AI-generated, not just the overall document.",
        "YES", "-", "Already implemented — ai_detector.py returns flagged_sentences list shown in teacher_ui.py"
    ),
    (
        "Burstiness Measurement",
        "Measures variation in sentence lengths. AI writing is uniform; human writing mixes short and long sentences naturally.",
        "PARTIAL", "-",
        "Burstiness is calculated in originality.py but used only as part of the composite originality score. Not surfaced as a standalone named metric to the teacher."
    ),
    (
        "Perplexity Scoring",
        "Measures how predictable each word choice is. AI always picks the most expected word (low perplexity). Humans pick unexpected personal words (high perplexity). Does not penalise students who write clearly.",
        "NO", "YES — Confidently",
        "Uses Ollama model already running. No new library needed. Directly solves the false positive problem where good human writers get wrongly flagged as AI."
    ),
    (
        "Mixed Content Detection",
        "Detects when only part of a submission is AI-generated — for example, 3 paragraphs written by student and 2 paragraphs generated by ChatGPT.",
        "NO", "YES — Partially",
        "Sentence-level AI flagging already exists. Aggregating flagged sentence count into a mixed-content percentage is an extension of existing logic."
    ),
    (
        "Humanized AI Detection",
        "Detects when a student used an AI tool specifically designed to rewrite ChatGPT output to sound more human (e.g. Undetectable.ai).",
        "NO", "NO",
        "Requires a model specifically trained on humanized AI samples. General purpose models like Qwen2.5 cannot reliably detect this. No local solution exists currently."
    ),
]

ORIGINALITY_AI = [
    (
        "AI Writing Detection",
        "Detects AI-generated content in student submissions.",
        "YES", "-", "Already implemented in ai_detector.py."
    ),
    (
        "Peer Plagiarism Detection",
        "Detects copying between students in the same batch.",
        "YES", "-", "Already implemented in copy_checker.py."
    ),
    (
        "Combined Integrity Score",
        "Single score combining all detection signals for easy teacher review.",
        "YES", "-", "Already implemented in report_builder.py."
    ),
    (
        "Readability Grade Level Check",
        "Measures writing complexity using the Flesch-Kincaid formula. Flags when a school student submits text written at PhD-level vocabulary — a strong indicator of AI use that bypasses style detection.",
        "NO", "YES — Confidently",
        "Pure mathematical formula. No AI model, no library, approximately 20 lines of Python. Zero false positive risk. A Class 10 student cannot accidentally write at Grade 16 vocabulary level."
    ),
    (
        "Ensemble Majority Voting",
        "Runs detection through multiple models and requires majority agreement before flagging. Reduces false positives dramatically — one detector being wrong cannot cause a false verdict.",
        "NO", "YES — Confidently",
        "Logic change only in report_builder.py. No new model or library. Directly reduces false positives by requiring 3 or more detectors to agree before issuing a high-confidence flag."
    ),
    (
        "Humanized AI Detection",
        "Specifically detects AI content that has been run through humanizing tools to bypass standard AI detectors.",
        "NO", "NO",
        "Requires a model trained specifically on humanized AI output paired with original AI output. Cannot be solved with a general purpose model like Qwen2.5. Specialized training data needed."
    ),
    (
        "Cross-Submission Historical Comparison",
        "Compares current submissions against all previously submitted assignments across past academic years.",
        "NO", "YES — Confidently",
        "The copy detection logic already exists. Requires storing past submissions in a local SQLite database (built into Python, no install needed) and including them in the copy checker run."
    ),
]

COPYLEAKS = [
    (
        "AI Writing Detection",
        "Detects AI-generated content using a custom trained model.",
        "YES", "-", "Already implemented using Qwen2.5 via Ollama."
    ),
    (
        "Peer Copy Detection",
        "Detects copying between students.",
        "YES", "-", "Already implemented using TF-IDF in copy_checker.py."
    ),
    (
        "Paraphrase Detection",
        "Detects meaning-level similarity even when words are different.",
        "YES", "-", "Already implemented using nomic-embed-text embeddings."
    ),
    (
        "Document Metadata Forensics",
        "Reads hidden timestamps inside Word/PDF files. If 1500 words were created and saved in 3 minutes — impossible for human typing — it is flagged as AI paste evidence. Independent of text content.",
        "NO", "YES — Confidently",
        "python-docx and PyMuPDF already installed in project. Metadata reading is 15 lines of code in text_extractor.py. Creates a detection signal that cannot be defeated by editing the text content."
    ),
    (
        "OCR for Scanned Documents",
        "Reads text from image-based PDFs and scanned handwritten assignments so they can be checked for plagiarism.",
        "NO", "YES — With Setup",
        "Requires installing Tesseract-OCR software on the computer and the pytesseract Python library. One additional code branch in text_extractor.py. Closes the gap where photograph submissions bypass all detection."
    ),
    (
        "Cross-Language Detection",
        "Detects when a student translated content from another language and submitted the English version.",
        "NO", "NO",
        "Requires a multilingual embedding model trained across 100+ languages. The nomic-embed-text model currently in use is English-only. Switching to a multilingual model is possible but accuracy cannot be guaranteed without testing."
    ),
    (
        "Humanized AI Detection",
        "Detects AI content specifically rewritten to avoid standard AI detectors.",
        "NO", "NO",
        "Requires specialized training data of humanized AI pairs. Not solvable with a general purpose local model. No reliable local solution currently available."
    ),
    (
        "Internet-Wide Plagiarism Check",
        "Checks submission against content published on the internet.",
        "NO", "NO",
        "Requires internet access and a crawled web index. Not feasible for an offline local system. Can be added as an optional API call to Copyleaks or similar service if internet access is available."
    ),
]

TOOL_SHEETS = [
    ("Turnitin", TURNITIN),
    ("GPTZero", GPTZERO),
    ("Originality.ai", ORIGINALITY_AI),
    ("Copyleaks", COPYLEAKS),
]

COL_WIDTHS = [28, 42, 14, 20, 52]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

for tool_name, rows in TOOL_SHEETS:
    ws = wb.create_sheet(title=tool_name)
    set_col_widths(ws, COL_WIDTHS)
    write_header(ws, f"{tool_name} — Feature Analysis", COLS)

    for i, (feat, what, status, can, reason) in enumerate(rows, start=3):
        display_can = can if can != "-" else "Already Present"
        display_reason = reason
        write_row(ws, i, [feat, what, status, display_can, display_reason], status)

    ws.freeze_panes = "A3"   # keep header visible while scrolling

wb.save(OUTPUT)
print(f"\n  Excel saved to:\n  {OUTPUT}\n")
print("  Sheets:", [s.title for s in wb.worksheets])
